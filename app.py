import os
import io
import secrets
from functools import wraps
from decimal import Decimal
from collections import Counter, defaultdict
from datetime import datetime, date, timedelta
from urllib.parse import urlparse, urljoin

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    session,
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from werkzeug.utils import secure_filename
from flask import send_from_directory

from flask_login import (
    LoginManager,
    UserMixin,
    login_user,
    logout_user,
    current_user,
)
from werkzeug.security import generate_password_hash, check_password_hash

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


# -----------------------
# App & DB Config
# -----------------------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key-change-me")

db_url = os.environ.get("DATABASE_URL", "sqlite:///crm.db")

# Railway/Heroku kadang pakai "postgres://", SQLAlchemy maunya "postgresql://"
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

# Paksa SQLAlchemy pakai psycopg v3 (bukan psycopg2)
if db_url.startswith("postgresql://"):
    db_url = db_url.replace("postgresql://", "postgresql+psycopg://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)


# -----------------------
# Upload Config
# -----------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

ALLOWED_EXT = {"pdf", "jpg", "jpeg", "png"}


# -----------------------
# Login Manager (username/password)
# -----------------------
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message_category = "error"


@login_manager.unauthorized_handler
def unauthorized():
    # Hindari redirect loop. Kalau belum login, selalu lempar ke /login?next=...
    next_url = request.full_path if request.query_string else request.path
    if next_url.startswith("/login"):
        next_url = "/"
    return redirect(url_for("login", next=next_url))


def is_safe_url(target: str) -> bool:
    # Cegah open redirect
    if not target:
        return False
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    return test_url.scheme in ("http", "https") and ref_url.netloc == test_url.netloc


# -----------------------
# Helpers - tenant
# -----------------------
def current_company_id():
    cid = session.get("company_id")
    if cid:
        try:
            return int(cid)
        except Exception:
            return None

    # fallback: kalau login via username/password, ambil dari user
    if getattr(current_user, "is_authenticated", False) and getattr(current_user, "company_id", None):
        session["company_id"] = int(current_user.company_id)
        return int(current_user.company_id)

    return None


def tenant_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_company_id():
            return redirect(url_for("tenant_login", next=request.path))
        return fn(*args, **kwargs)
    return wrapper


def require_company_query(model):
    """Helper query scoped per company_id"""
    cid = current_company_id()
    return model.query.filter_by(company_id=cid)

# -----------------------
# Admin helpers (buat tenant)
# -----------------------
def is_admin_user() -> bool:
    # Admin = username sama dengan env ADMIN_USERNAME (default: admin)
    admin_username = os.environ.get("ADMIN_USERNAME", "admin")
    return current_user.is_authenticated and (current_user.username == admin_username)

def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not is_admin_user():
            abort(403)
        return fn(*args, **kwargs)
    return wrapper

def generate_access_code() -> str:
    # kode pendek, uppercase, gampang diketik
    # contoh: BP-8F3K2Q
    while True:
        code = "BP-" + secrets.token_hex(3).upper()  # 6 hex
        if not Company.query.filter_by(access_code=code).first():
            return code

def normalize_pin(pin: str) -> str:
    return (pin or "").strip()

# -----------------------
# Admin Routes - Tenants
# -----------------------
@app.get("/admin/tenants")
@login_required
@admin_required
def admin_tenants_list():
    tenants = Company.query.order_by(Company.created_at.desc()).all()
    return render_template("tenants_list.html", tenants=tenants)

@app.get("/admin/tenants/new")
@login_required
@admin_required
def admin_tenants_new():
    # akses_code default auto
    default_code = generate_access_code()
    return render_template("tenant_new.html", default_code=default_code)

@app.post("/admin/tenants/new")
@login_required
@admin_required
def admin_tenants_create():
    name = (request.form.get("name") or "").strip()
    access_code = (request.form.get("access_code") or "").strip().upper()
    pin = normalize_pin(request.form.get("pin"))

    if not name:
        flash("Nama tenant/perusahaan wajib diisi.", "error")
        return redirect(url_for("admin_tenants_new"))

    if not access_code:
        access_code = generate_access_code()

    if Company.query.filter_by(access_code=access_code).first():
        flash("Access Code sudah dipakai. Coba generate ulang.", "error")
        return redirect(url_for("admin_tenants_new"))

    # PIN minimal 4 digit/karakter biar aman
    if not pin or len(pin) < 4:
        flash("PIN minimal 4 karakter/angka.", "error")
        return redirect(url_for("admin_tenants_new"))

    c = Company(name=name, access_code=access_code, pin_hash="temp")
    c.set_pin(pin)
    db.session.add(c)
    db.session.commit()

    flash(f"Tenant dibuat: {c.name} (Access Code: {c.access_code})", "success")
    return redirect(url_for("admin_tenants_list"))


# -----------------------
# Models
# -----------------------
class Company(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(160), nullable=False)
    access_code = db.Column(db.String(40), unique=True, nullable=False, index=True)
    pin_hash = db.Column(db.String(255), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    def set_pin(self, pin: str):
        self.pin_hash = generate_password_hash(pin)

    def check_pin(self, pin: str) -> bool:
        return check_password_hash(self.pin_hash, pin)


class LeadSource(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=False, index=True)
    name = db.Column(db.String(120), nullable=False)

    company = db.relationship("Company")

    __table_args__ = (
        db.UniqueConstraint("company_id", "name", name="uq_lead_source_company_name"),
    )


class Need(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=False, index=True)
    name = db.Column(db.String(120), nullable=False)

    company = db.relationship("Company")

    __table_args__ = (
        db.UniqueConstraint("company_id", "name", name="uq_need_company_name"),
    )


class Progress(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=False, index=True)
    name = db.Column(db.String(120), nullable=False)

    company = db.relationship("Company")

    __table_args__ = (
        db.UniqueConstraint("company_id", "name", name="uq_progress_company_name"),
    )


class FollowUpStage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=False, index=True)
    name = db.Column(db.String(120), nullable=False)

    company = db.relationship("Company")

    __table_args__ = (
        db.UniqueConstraint("company_id", "name", name="uq_stage_company_name"),
    )


class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=False, index=True)

    name = db.Column(db.String(160), nullable=False)
    salesman_name = db.Column(db.String(160), nullable=True)

    address = db.Column(db.Text, nullable=True)
    phone_wa = db.Column(db.String(80), nullable=True)
    email = db.Column(db.String(160), nullable=True)
    pic = db.Column(db.String(160), nullable=True)

    lead_source_id = db.Column(db.Integer, db.ForeignKey("lead_source.id"), nullable=True)
    need_id = db.Column(db.Integer, db.ForeignKey("need.id"), nullable=True)
    progress_id = db.Column(db.Integer, db.ForeignKey("progress.id"), nullable=True)

    prospect_date = db.Column(db.Date, nullable=True)
    estimated_value = db.Column(db.Numeric(14, 2), nullable=True)
    prospect_next_followup_date = db.Column(db.Date, nullable=True)

    note_followup_awal = db.Column(db.Text, nullable=True)
    note_followup_lanjutan = db.Column(db.Text, nullable=True)
    management_comment = db.Column(db.Text, nullable=True)

    lead_source = db.relationship("LeadSource")
    need = db.relationship("Need")
    progress = db.relationship("Progress")

    followups = db.relationship("FollowUpLog", backref="customer", cascade="all, delete-orphan")


class FollowUpLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=False, index=True)

    customer_id = db.Column(db.Integer, db.ForeignKey("customer.id"), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    stage_id = db.Column(db.Integer, db.ForeignKey("follow_up_stage.id"), nullable=True)
    note = db.Column(db.Text, nullable=True)

    stage = db.relationship("FollowUpStage")


class Attachment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=False, index=True)

    customer_id = db.Column(db.Integer, db.ForeignKey("customer.id"), nullable=False)

    original_filename = db.Column(db.String(255), nullable=False)
    stored_filename = db.Column(db.String(255), nullable=False)
    mime_type = db.Column(db.String(120), nullable=True)
    size_bytes = db.Column(db.Integer, nullable=True)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    customer = db.relationship("Customer", backref=db.backref("attachments", cascade="all, delete-orphan"))


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey("company.id"), nullable=False, index=True)

    # NOTE: tetap unique global biar gak perlu migrasi berat (kalau sudah terlanjur)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)

    def set_password(self, password: str):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password: str) -> bool:
        return check_password_hash(self.password_hash, password)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# -----------------------
# Master map
# -----------------------
MASTER_MAP = {
    "sources": (LeadSource, "Sumber Prospek"),
    "needs": (Need, "Produk/Jasa Dibutuhkan"),
    "progress": (Progress, "Progress"),
    "stages": (FollowUpStage, "Tahap Progress Follow Up"),
}


# -----------------------
# Helpers - parsing / formatting
# -----------------------
def to_int_or_none(v):
    try:
        if v is None or str(v).strip() == "":
            return None
        return int(v)
    except Exception:
        return None


def to_date_or_none(v: str):
    try:
        v = (v or "").strip()
        if not v:
            return None
        return datetime.strptime(v, "%Y-%m-%d").date()
    except Exception:
        return None


def parse_date_yyyy_mm_dd(s: str):
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def to_decimal_or_none(v: str):
    try:
        v = (v or "").strip()
        if not v:
            return None
        v = v.replace(".", "").replace(",", "").replace("Rp", "").strip()
        return Decimal(v)
    except Exception:
        return None


def to_decimal_or_zero(v):
    try:
        if v is None or v == "":
            return Decimal("0")
        return Decimal(str(v))
    except Exception:
        return Decimal("0")


def fmt_idr(x):
    try:
        if x is None:
            return "-"
        n = int(Decimal(x))
        return "Rp {:,}".format(n).replace(",", ".")
    except Exception:
        return "-"


def is_won_progress(name: str) -> bool:
    if not name:
        return False
    p = name.lower()
    return any(k in p for k in ["won", "closing", "closed", "deal", "berhasil", "success"])


def is_lost_progress(name: str) -> bool:
    if not name:
        return False
    p = name.lower()
    return any(k in p for k in ["lost", "gagal", "failed", "cancel", "batal"])


def is_offer_progress(name: str) -> bool:
    if not name:
        return False
    p = name.lower()
    return any(k in p for k in ["proposal", "penawaran", "quotation", "quote", "offer"])


def progress_flag(progress_name: str):
    p = (progress_name or "").lower()
    is_won = any(k in p for k in ["won", "closing", "closed", "deal", "berhasil", "success"])
    is_lost = any(k in p for k in ["lost", "gagal", "failed", "cancel", "batal"])
    return is_won, is_lost


def week_range(today: date):
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday


# -----------------------
# Schema helper (minimal, no migration tool)
# -----------------------
def ensure_schema():
    """
    Minimal ALTER untuk menambah kolom yang mungkin belum ada.
    Fokus: company_id + prospect_next_followup_date (biar gak crash).
    """
    engine = db.engine
    dialect = engine.dialect.name  # "sqlite" / "postgresql"

    def col_exists_pg(conn, table: str, col: str) -> bool:
        q = """
        SELECT 1
        FROM information_schema.columns
        WHERE table_schema='public'
          AND table_name=%s
          AND column_name=%s
        LIMIT 1
        """
        return conn.exec_driver_sql(q, (table, col)).first() is not None

    def ensure_col(table: str, col: str, ddl_sqlite: str, ddl_pg: str):
        with engine.begin() as conn:
            if dialect == "sqlite":
                cols = [r[1] for r in conn.exec_driver_sql(f"PRAGMA table_info({table})").fetchall()]
                if col not in cols:
                    conn.exec_driver_sql(ddl_sqlite)
            else:
                if not col_exists_pg(conn, table, col):
                    conn.exec_driver_sql(ddl_pg)

    # customer.prospect_next_followup_date
    ensure_col(
        "customer",
        "prospect_next_followup_date",
        "ALTER TABLE customer ADD COLUMN prospect_next_followup_date DATE",
        "ALTER TABLE customer ADD COLUMN prospect_next_followup_date DATE",
    )

    # company_id columns (kalau ada tabel lama yang belum punya)
    ensure_col(
        "customer",
        "company_id",
        "ALTER TABLE customer ADD COLUMN company_id INTEGER",
        "ALTER TABLE customer ADD COLUMN company_id INTEGER",
    )
    ensure_col(
        "lead_source",
        "company_id",
        "ALTER TABLE lead_source ADD COLUMN company_id INTEGER",
        "ALTER TABLE lead_source ADD COLUMN company_id INTEGER",
    )
    ensure_col(
        "need",
        "company_id",
        "ALTER TABLE need ADD COLUMN company_id INTEGER",
        "ALTER TABLE need ADD COLUMN company_id INTEGER",
    )
    ensure_col(
        "progress",
        "company_id",
        "ALTER TABLE progress ADD COLUMN company_id INTEGER",
        "ALTER TABLE progress ADD COLUMN company_id INTEGER",
    )
    ensure_col(
        "follow_up_stage",
        "company_id",
        "ALTER TABLE follow_up_stage ADD COLUMN company_id INTEGER",
        "ALTER TABLE follow_up_stage ADD COLUMN company_id INTEGER",
    )
    ensure_col(
        "follow_up_log",
        "company_id",
        "ALTER TABLE follow_up_log ADD COLUMN company_id INTEGER",
        "ALTER TABLE follow_up_log ADD COLUMN company_id INTEGER",
    )
    ensure_col(
        "attachment",
        "company_id",
        "ALTER TABLE attachment ADD COLUMN company_id INTEGER",
        "ALTER TABLE attachment ADD COLUMN company_id INTEGER",
    )
    ensure_col(
        "user",
        "company_id",
        "ALTER TABLE user ADD COLUMN company_id INTEGER",
        "ALTER TABLE \"user\" ADD COLUMN company_id INTEGER",
    )

    # Default company_id = 1 untuk data lama (biar tidak undefined)
    with engine.begin() as conn:
        if dialect == "sqlite":
            conn.exec_driver_sql("UPDATE customer SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE lead_source SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE need SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE progress SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE follow_up_stage SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE follow_up_log SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE attachment SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE user SET company_id = 1 WHERE company_id IS NULL")
        else:
            conn.exec_driver_sql("UPDATE customer SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE lead_source SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE need SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE progress SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE follow_up_stage SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE follow_up_log SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE attachment SET company_id = 1 WHERE company_id IS NULL")
            conn.exec_driver_sql("UPDATE \"user\" SET company_id = 1 WHERE company_id IS NULL")


# -----------------------
# Seed data (per company)
# -----------------------
def ensure_seed_data():
    # buat 1 company default kalau belum ada
    if Company.query.count() == 0:
        code = os.environ.get("DEFAULT_ACCESS_CODE") or secrets.token_hex(4)
        pin = os.environ.get("DEFAULT_PIN") or "123456"
        comp = Company(name="Default Company", access_code=code)
        comp.set_pin(pin)
        db.session.add(comp)
        db.session.commit()
        print(f"[seed] created company access_code={code} pin={pin}")

    default_company = Company.query.order_by(Company.id.asc()).first()
    default_cid = default_company.id

    # Seed tahap default (PER COMPANY)
    if FollowUpStage.query.filter_by(company_id=default_cid).count() == 0:
        defaults = ["New", "Contacted", "Qualified", "Proposal", "Negotiation", "Won", "Lost"]
        for n in defaults:
            db.session.add(FollowUpStage(company_id=default_cid, name=n))
        db.session.commit()

    # Seed progress default (PER COMPANY) - optional tapi membantu
    if Progress.query.filter_by(company_id=default_cid).count() == 0:
        defaults = ["New", "Contacted", "Proposal", "Negotiation", "Won", "Lost"]
        for n in defaults:
            db.session.add(Progress(company_id=default_cid, name=n))
        db.session.commit()

    # Seed admin user default kalau belum ada user
    if User.query.count() == 0:
        admin_user = os.environ.get("ADMIN_USERNAME", "admin")
        admin_pass = os.environ.get("ADMIN_PASSWORD", "admin12345")
        u = User(username=admin_user, company_id=default_cid)
        u.set_password(admin_pass)
        db.session.add(u)
        db.session.commit()
        print(f"[seed] created admin user: {admin_user}")


# -----------------------
# Context helpers to templates
# -----------------------
@app.context_processor
def inject_helpers():
    return dict(fmt_idr=fmt_idr)

@app.context_processor
def inject_helpers():
    # tenant info
    cid = session.get("company_id")
    company_name = None
    if cid:
        c = Company.query.get(int(cid))
        company_name = c.name if c else None

    return dict(
        fmt_idr=fmt_idr,
        current_company_name=company_name,
        current_company_id=cid,
        is_admin=is_admin_user() if 'is_admin_user' in globals() else False
    )


# -----------------------
# Auth Routes (username/password)
# -----------------------
@app.get("/login")
def login():
    if current_user.is_authenticated:
        # sync company_id session
        current_company_id()
        return redirect(url_for("home"))
    return render_template("login.html")


@app.post("/login")
def login_post():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")

    u = User.query.filter_by(username=username).first()
    if not u or not u.check_password(password):
        flash("Username / password salah.", "error")
        return redirect(url_for("login"))

    login_user(u)
    session["company_id"] = int(u.company_id)  # penting untuk tenant scope
    flash("Login berhasil.", "success")

    next_url = request.args.get("next")
    if next_url and is_safe_url(next_url):
        return redirect(next_url)
    return redirect(url_for("home"))


@app.get("/logout")
@tenant_required
def logout():
    logout_user()
    session.pop("company_id", None)
    flash("Logout berhasil.", "success")
    return redirect(url_for("login"))


# -----------------------
# Tenant Auth Routes (access code + pin)
# -----------------------
@app.get("/tenant-login")
def tenant_login():
    if current_company_id():
        return redirect(url_for("home"))
    return render_template("tenant_login.html")


@app.post("/tenant-login")
def tenant_login_post():
    access_code = request.form.get("access_code", "").strip()
    pin = request.form.get("pin", "").strip()

    c = Company.query.filter_by(access_code=access_code).first()
    if not c or not c.check_pin(pin):
        flash("Access Code / PIN salah.", "error")
        return redirect(url_for("tenant_login"))

    session["company_id"] = int(c.id)
    flash(f"Login berhasil: {c.name}", "success")

    next_url = request.args.get("next")
    return redirect(next_url or url_for("home"))


@app.get("/tenant-logout")
def tenant_logout():
    session.pop("company_id", None)
    flash("Logout berhasil.", "success")
    return redirect(url_for("tenant_login"))


# -----------------------
# Routes - Dashboard
# -----------------------
@app.get("/")
@tenant_required
def home():
    cid = current_company_id()

    customers_count = Customer.query.filter_by(company_id=cid).count()
    latest_followups = (
        FollowUpLog.query
        .filter_by(company_id=cid)
        .order_by(FollowUpLog.created_at.desc())
        .limit(10)
        .all()
    )

    rows = (
        db.session.query(
            Customer.id,
            Customer.salesman_name,
            Customer.estimated_value,
            Progress.name
        )
        .outerjoin(Progress, Customer.progress_id == Progress.id)
        .filter(Customer.company_id == cid)
        .all()
    )

    total = 0
    won_count = 0
    lost_count = 0
    won_value = Decimal("0")
    lost_value = Decimal("0")
    sales_won = {}

    for _id, salesman, est, prog_name in rows:
        total += 1
        is_won, is_lost = progress_flag(prog_name)

        if is_won:
            won_count += 1
            if est:
                won_value += Decimal(str(est))
            key = (salesman or "Unknown").strip() or "Unknown"
            sales_won[key] = sales_won.get(key, 0) + 1
        elif is_lost:
            lost_count += 1
            if est:
                lost_value += Decimal(str(est))

    other_count = max(total - won_count - lost_count, 0)
    not_won = max(total - won_count, 0)
    not_lost = max(total - lost_count, 0)

    # ===== Estimasi Nilai: Sudah Penawaran vs Belum Penawaran =====
    all_customers = Customer.query.filter_by(company_id=cid).all()
    offer_value = Decimal("0")
    not_offer_value = Decimal("0")

    for c in all_customers:
        val = to_decimal_or_zero(c.estimated_value)
        pname = (c.progress.name if c.progress else "")
        if is_offer_progress(pname):
            offer_value += val
        else:
            not_offer_value += val

    pie_offer_labels = ["Sudah Penawaran", "Belum Penawaran"]
    pie_offer_values = [float(offer_value), float(not_offer_value)]

    # Top sales
    top_sales = sorted(sales_won.items(), key=lambda x: x[1], reverse=True)[:10]
    top_sales_labels = [x[0] for x in top_sales]
    top_sales_values = [x[1] for x in top_sales]

    dashboard = {
        "total": total,
        "won_count": won_count,
        "lost_count": lost_count,
        "other_count": other_count,
        "not_won": not_won,
        "not_lost": not_lost,
        "won_value": int(won_value),
        "lost_value": int(lost_value),
        "top_sales_labels": top_sales_labels,
        "top_sales_values": top_sales_values,
    }

    # ===== PIE: sumber prospek =====
    source_rows = (
        db.session.query(LeadSource.name, func.count(Customer.id))
        .join(Customer, Customer.lead_source_id == LeadSource.id)
        .filter(LeadSource.company_id == cid)
        .filter(Customer.company_id == cid)
        .group_by(LeadSource.name)
        .order_by(func.count(Customer.id).desc())
        .all()
    )
    pie_sources_labels = [r[0] for r in source_rows]
    pie_sources_values = [int(r[1]) for r in source_rows]

    # ===== PIE: produk/jasa (need) =====
    need_rows = (
        db.session.query(Need.name, func.count(Customer.id))
        .join(Customer, Customer.need_id == Need.id)
        .filter(Need.company_id == cid)
        .filter(Customer.company_id == cid)
        .group_by(Need.name)
        .order_by(func.count(Customer.id).desc())
        .all()
    )
    pie_needs_labels = [r[0] for r in need_rows]
    pie_needs_values = [int(r[1]) for r in need_rows]

    # Reminder followup minggu ini
    today = date.today()
    d1, d2 = week_range(today)
    followup_week = (
        Customer.query
        .filter_by(company_id=cid)
        .filter(Customer.prospect_next_followup_date.isnot(None))
        .filter(Customer.prospect_next_followup_date >= d1)
        .filter(Customer.prospect_next_followup_date <= d2)
        .order_by(Customer.prospect_next_followup_date.asc())
        .limit(50)
        .all()
    )

    # Top sumber prospek yg menghasilkan closing (count + total nilai)
    won_customers = (
        Customer.query
        .filter_by(company_id=cid)
        .join(Customer.progress, isouter=True)
        .join(Customer.lead_source, isouter=True)
        .all()
    )

    source_won_count = Counter()
    source_won_value = defaultdict(Decimal)

    for c in won_customers:
        pname = c.progress.name if getattr(c, "progress", None) else ""
        if is_won_progress(pname):
            sname = c.lead_source.name if c.lead_source else "(Tanpa Sumber)"
            source_won_count[sname] += 1
            source_won_value[sname] += to_decimal_or_zero(getattr(c, "estimated_value", None))

    top_sources = sorted(source_won_count.items(), key=lambda x: x[1], reverse=True)[:10]
    top_sources_labels = [k for k, _ in top_sources]
    top_sources_values = [v for _, v in top_sources]
    top_sources_value_sum = [float(source_won_value[k]) for k in top_sources_labels]

    return render_template(
        "home.html",
        customers_count=customers_count,
        latest_followups=latest_followups,
        followup_week=followup_week,
        dashboard=dashboard,
        pie_sources_labels=pie_sources_labels,
        pie_sources_values=pie_sources_values,
        pie_needs_labels=pie_needs_labels,
        pie_needs_values=pie_needs_values,
        pie_offer_labels=pie_offer_labels,
        pie_offer_values=pie_offer_values,
        week_start=d1,
        week_end=d2,
        top_sources_won_labels=top_sources_labels,
        top_sources_won_values=top_sources_values,
        top_sources_won_value_sum=top_sources_value_sum,
    )


# -----------------------
# Routes - Customers
# -----------------------
@app.get("/customers")
@tenant_required
def customers_list():
    cid = current_company_id()

    q = (request.args.get("q", "") or "").strip()
    date_from = (request.args.get("date_from", "") or "").strip()
    date_to = (request.args.get("date_to", "") or "").strip()

    query = Customer.query.filter_by(company_id=cid)

    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                Customer.name.ilike(like),
                Customer.pic.ilike(like),
                Customer.phone_wa.ilike(like),
                Customer.email.ilike(like),
            )
        )

    df = to_date_or_none(date_from)
    dt = to_date_or_none(date_to)
    if df:
        query = query.filter(Customer.prospect_date >= df)
    if dt:
        query = query.filter(Customer.prospect_date <= dt)

    customers = query.order_by(Customer.id.desc()).all()

    return render_template(
        "customers_list.html",
        customers=customers,
        q=q,
        date_from=date_from,
        date_to=date_to,
    )


@app.get("/customers/new")
@tenant_required
def customers_new():
    return render_customer_form(Customer(), is_edit=False)


@app.post("/customers/new")
@tenant_required
def customers_create():
    cid = current_company_id()

    c = Customer(
        company_id=cid,
        name=request.form.get("name", "").strip(),
        salesman_name=request.form.get("salesman_name", "").strip(),
        address=request.form.get("address", "").strip(),
        phone_wa=request.form.get("phone_wa", "").strip(),
        email=request.form.get("email", "").strip(),
        pic=request.form.get("pic", "").strip(),
        lead_source_id=to_int_or_none(request.form.get("lead_source_id")),
        need_id=to_int_or_none(request.form.get("need_id")),
        progress_id=to_int_or_none(request.form.get("progress_id")),
        prospect_date=to_date_or_none(request.form.get("prospect_date")),
        estimated_value=to_decimal_or_none(request.form.get("estimated_value")),
        note_followup_awal=request.form.get("note_followup_awal", "").strip(),
        note_followup_lanjutan=request.form.get("note_followup_lanjutan", "").strip(),
        management_comment=request.form.get("management_comment", "").strip(),
        prospect_next_followup_date=to_date_or_none(request.form.get("prospect_next_followup_date")),
    )

    if not c.name:
        flash("Nama customer wajib diisi.", "error")
        return render_customer_form(c, is_edit=False)

    db.session.add(c)
    db.session.commit()
    flash("Customer dibuat.", "success")
    return redirect(url_for("customer_detail", customer_id=c.id))


@app.get("/customers/<int:customer_id>")
@tenant_required
def customer_detail(customer_id: int):
    cid = current_company_id()

    c = Customer.query.filter_by(company_id=cid, id=customer_id).first_or_404()

    stages = FollowUpStage.query.filter_by(company_id=cid).order_by(FollowUpStage.name.asc()).all()

    followups = (
        FollowUpLog.query
        .filter_by(company_id=cid, customer_id=c.id)
        .order_by(FollowUpLog.created_at.desc())
        .all()
    )

    return render_template("customer_detail.html", c=c, stages=stages, followups=followups)


@app.get("/customers/<int:customer_id>/edit")
@tenant_required
def customers_edit(customer_id: int):
    cid = current_company_id()
    c = Customer.query.filter_by(company_id=cid, id=customer_id).first_or_404()
    return render_customer_form(c, is_edit=True)


@app.post("/customers/<int:customer_id>/edit")
@tenant_required
def customers_update(customer_id: int):
    cid = current_company_id()
    c = Customer.query.filter_by(company_id=cid, id=customer_id).first_or_404()

    c.name = request.form.get("name", "").strip()
    c.salesman_name = request.form.get("salesman_name", "").strip()
    c.prospect_date = to_date_or_none(request.form.get("prospect_date"))
    c.address = request.form.get("address", "").strip()
    c.phone_wa = request.form.get("phone_wa", "").strip()
    c.email = request.form.get("email", "").strip()
    c.pic = request.form.get("pic", "").strip()
    c.lead_source_id = to_int_or_none(request.form.get("lead_source_id"))
    c.need_id = to_int_or_none(request.form.get("need_id"))
    c.estimated_value = to_decimal_or_none(request.form.get("estimated_value"))
    c.progress_id = to_int_or_none(request.form.get("progress_id"))

    c.note_followup_awal = request.form.get("note_followup_awal", "").strip()
    c.note_followup_lanjutan = request.form.get("note_followup_lanjutan", "").strip()
    c.management_comment = request.form.get("management_comment", "").strip()

    c.prospect_next_followup_date = to_date_or_none(request.form.get("prospect_next_followup_date"))

    if not c.name:
        flash("Nama customer wajib diisi.", "error")
        return render_customer_form(c, is_edit=True)

    db.session.commit()
    flash("Customer diupdate.", "success")
    return redirect(url_for("customer_detail", customer_id=c.id))


@app.post("/customers/<int:customer_id>/delete")
@tenant_required
def customers_delete(customer_id: int):
    cid = current_company_id()
    c = Customer.query.filter_by(company_id=cid, id=customer_id).first_or_404()
    db.session.delete(c)
    db.session.commit()
    flash("Customer dihapus.", "success")
    return redirect(url_for("customers_list"))


def render_customer_form(c: Customer, is_edit: bool):
    cid = current_company_id()
    sources = LeadSource.query.filter_by(company_id=cid).order_by(LeadSource.name.asc()).all()
    needs = Need.query.filter_by(company_id=cid).order_by(Need.name.asc()).all()
    progresses = Progress.query.filter_by(company_id=cid).order_by(Progress.name.asc()).all()

    return render_template(
        "customer_form.html",
        c=c,
        is_edit=is_edit,
        sources=sources,
        needs=needs,
        progresses=progresses,
    )


# -----------------------
# Attachments
# -----------------------
@app.post("/customers/<int:customer_id>/attachments")
@tenant_required
def attachments_upload(customer_id: int):
    cid = current_company_id()
    c = Customer.query.filter_by(company_id=cid, id=customer_id).first_or_404()

    f = request.files.get("file")
    if not f or f.filename.strip() == "":
        flash("Pilih file dulu.", "error")
        return redirect(url_for("customer_detail", customer_id=c.id))

    filename = secure_filename(f.filename)
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ALLOWED_EXT:
        flash("File harus PDF/JPG/PNG.", "error")
        return redirect(url_for("customer_detail", customer_id=c.id))

    ts = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
    stored = f"{cid}_{c.id}_{ts}_{filename}"
    path = os.path.join(UPLOAD_DIR, stored)

    f.save(path)
    size = os.path.getsize(path)

    att = Attachment(
        company_id=cid,
        customer_id=c.id,
        original_filename=filename,
        stored_filename=stored,
        mime_type=f.mimetype,
        size_bytes=size,
    )
    db.session.add(att)
    db.session.commit()

    flash("Lampiran berhasil diupload.", "success")
    return redirect(url_for("customer_detail", customer_id=c.id))


@app.get("/attachments/<int:att_id>/download")
@tenant_required
def attachments_download(att_id: int):
    cid = current_company_id()
    att = Attachment.query.filter_by(company_id=cid, id=att_id).first_or_404()

    return send_from_directory(
        UPLOAD_DIR,
        att.stored_filename,
        as_attachment=True,
        download_name=att.original_filename,
    )


@app.post("/attachments/<int:att_id>/delete")
@tenant_required
def attachments_delete(att_id: int):
    cid = current_company_id()
    att = Attachment.query.filter_by(company_id=cid, id=att_id).first_or_404()

    cust_id = att.customer_id

    try:
        os.remove(os.path.join(UPLOAD_DIR, att.stored_filename))
    except Exception:
        pass

    db.session.delete(att)
    db.session.commit()
    flash("Lampiran dihapus.", "success")
    return redirect(url_for("customer_detail", customer_id=cust_id))


# -----------------------
# Import Customers (Excel)
# -----------------------
@app.get("/customers/import")
@tenant_required
def customers_import_page():
    return render_template("customers_import.html")


@app.post("/customers/import")
@tenant_required
def customers_import_post():
    cid = current_company_id()

    f = request.files.get("file")
    if not f or f.filename.strip() == "":
        flash("Pilih file Excel (.xlsx) dulu.", "error")
        return redirect(url_for("customers_import_page"))

    if not f.filename.lower().endswith(".xlsx"):
        flash("File harus .xlsx", "error")
        return redirect(url_for("customers_import_page"))

    bio = io.BytesIO(f.read())
    wb = load_workbook(bio)
    ws = wb.active

    header_row = [str(c.value or "").strip() for c in ws[1]]
    header_map = {h.lower(): i for i, h in enumerate(header_row)}

    def get_cell(row, key):
        idx = header_map.get(key.lower())
        if idx is None:
            return ""
        v = row[idx].value
        return "" if v is None else str(v).strip()

    created = 0
    updated = 0

    for r in ws.iter_rows(min_row=2):
        name = get_cell(r, "Nama")
        if not name:
            continue

        email = get_cell(r, "Email")
        phone = get_cell(r, "Telp/WA") or get_cell(r, "WA/Telp")
        salesman = get_cell(r, "Salesman")
        pic = get_cell(r, "PIC")
        address = get_cell(r, "Alamat")
        source_name = get_cell(r, "Sumber Prospek")
        need_name = get_cell(r, "Produk/Jasa") or get_cell(r, "Produk/Jasa Dibutuhkan")
        progress_name = get_cell(r, "Progress")
        est_value = get_cell(r, "Estimasi Nilai")
        prospect_date = get_cell(r, "Tgl Prospek")
        fu_next = get_cell(r, "Rencana FU") or get_cell(r, "Next Follow Up") or get_cell(r, "Rencana Follow Up")
        fu_awal = get_cell(r, "FU Awal") or get_cell(r, "Catatan Follow Up Awal")
        fu_lanjutan = get_cell(r, "FU Lanjutan") or get_cell(r, "Catatan Follow Up Lanjutan")
        km = get_cell(r, "Komen Manajemen")

        lead_source_id = None
        if source_name:
            ls = LeadSource.query.filter_by(company_id=cid, name=source_name).first()
            if not ls:
                ls = LeadSource(company_id=cid, name=source_name)
                db.session.add(ls)
                db.session.flush()
            lead_source_id = ls.id

        need_id = None
        if need_name:
            nd = Need.query.filter_by(company_id=cid, name=need_name).first()
            if not nd:
                nd = Need(company_id=cid, name=need_name)
                db.session.add(nd)
                db.session.flush()
            need_id = nd.id

        progress_id = None
        if progress_name:
            pg = Progress.query.filter_by(company_id=cid, name=progress_name).first()
            if not pg:
                pg = Progress(company_id=cid, name=progress_name)
                db.session.add(pg)
                db.session.flush()
            progress_id = pg.id

        # cari existing by email/phone (DALAM company ini saja)
        c = None
        if email:
            c = Customer.query.filter_by(company_id=cid, email=email).first()
        if not c and phone:
            c = Customer.query.filter_by(company_id=cid, phone_wa=phone).first()

        pd = parse_date_yyyy_mm_dd(prospect_date)
        ndt = parse_date_yyyy_mm_dd(fu_next)

        ev = None
        if est_value:
            cleaned = (
                est_value.replace("Rp", "")
                .replace(".", "")
                .replace(",", "")
                .strip()
            )
            try:
                ev = Decimal(cleaned)
            except Exception:
                ev = None

        if c:
            c.name = name
            c.salesman_name = salesman
            c.pic = pic
            c.address = address
            c.phone_wa = phone
            c.email = email
            c.lead_source_id = lead_source_id
            c.need_id = need_id
            c.progress_id = progress_id
            c.estimated_value = ev
            c.prospect_date = pd
            c.prospect_next_followup_date = ndt
            c.note_followup_awal = fu_awal
            c.note_followup_lanjutan = fu_lanjutan
            c.management_comment = km
            updated += 1
        else:
            c = Customer(
                company_id=cid,
                name=name,
                salesman_name=salesman,
                pic=pic,
                address=address,
                phone_wa=phone,
                email=email,
                lead_source_id=lead_source_id,
                need_id=need_id,
                progress_id=progress_id,
                estimated_value=ev,
                prospect_date=pd,
                prospect_next_followup_date=ndt,
                note_followup_awal=fu_awal,
                note_followup_lanjutan=fu_lanjutan,
                management_comment=km,
            )
            db.session.add(c)
            created += 1

    db.session.commit()
    flash(f"Import selesai. Created: {created}, Updated: {updated}", "success")
    return redirect(url_for("customers_list"))


# -----------------------
# Followup Log
# -----------------------
@app.post("/customers/<int:customer_id>/followups")
@tenant_required
def followups_add(customer_id: int):
    cid = current_company_id()
    c = Customer.query.filter_by(company_id=cid, id=customer_id).first_or_404()

    stage_id = to_int_or_none(request.form.get("stage_id"))
    note = request.form.get("note", "").strip()

    fu = FollowUpLog(company_id=cid, customer_id=c.id, stage_id=stage_id, note=note)
    db.session.add(fu)
    db.session.commit()
    flash("Follow up dicatat.", "success")
    return redirect(url_for("customer_detail", customer_id=c.id))


@app.post("/followups/<int:followup_id>/delete")
@tenant_required
def followups_delete(followup_id: int):
    cid = current_company_id()
    fu = FollowUpLog.query.filter_by(company_id=cid, id=followup_id).first_or_404()

    customer_id = fu.customer_id
    db.session.delete(fu)
    db.session.commit()
    flash("Log follow up dihapus.", "success")
    return redirect(url_for("customer_detail", customer_id=customer_id))


# -----------------------
# Masters CRUD
# -----------------------
def get_master_model(key: str):
    key = (key or "").strip().lower()
    if key not in MASTER_MAP:
        flash("Master tidak dikenal.", "error")
        return MASTER_MAP["sources"]
    return MASTER_MAP[key]


@app.get("/masters/<string:key>")
@tenant_required
def masters_list(key: str):
    cid = current_company_id()
    model, title = get_master_model(key)
    items = model.query.filter_by(company_id=cid).order_by(model.name.asc()).all()
    return render_template("masters_list.html", key=key, title=title, items=items)


@app.get("/masters/<string:key>/new")
@tenant_required
def masters_new(key: str):
    _, title = get_master_model(key)
    return render_template("master_form.html", key=key, title=title, item=None, is_edit=False)


@app.post("/masters/<string:key>/new")
@tenant_required
def masters_create(key: str):
    cid = current_company_id()
    model, title = get_master_model(key)
    name = request.form.get("name", "").strip()

    if not name:
        flash("Nama wajib diisi.", "error")
        return render_template("master_form.html", key=key, title=title, item=None, is_edit=False)

    if model.query.filter_by(company_id=cid, name=name).first():
        flash("Nama sudah ada.", "error")
        return render_template("master_form.html", key=key, title=title, item=None, is_edit=False)

    item = model(company_id=cid, name=name)
    db.session.add(item)
    db.session.commit()
    flash("Data master ditambahkan.", "success")
    return redirect(url_for("masters_list", key=key))


@app.get("/masters/<string:key>/<int:item_id>/edit")
@tenant_required
def masters_edit(key: str, item_id: int):
    cid = current_company_id()
    model, title = get_master_model(key)
    item = model.query.filter_by(company_id=cid, id=item_id).first_or_404()
    return render_template("master_form.html", key=key, title=title, item=item, is_edit=True)


@app.post("/masters/<string:key>/<int:item_id>/edit")
@tenant_required
def masters_update(key: str, item_id: int):
    cid = current_company_id()
    model, title = get_master_model(key)
    item = model.query.filter_by(company_id=cid, id=item_id).first_or_404()

    name = request.form.get("name", "").strip()
    if not name:
        flash("Nama wajib diisi.", "error")
        return render_template("master_form.html", key=key, title=title, item=item, is_edit=True)

    exists = model.query.filter(
        model.company_id == cid,
        model.name == name,
        model.id != item.id
    ).first()
    if exists:
        flash("Nama sudah dipakai item lain.", "error")
        return render_template("master_form.html", key=key, title=title, item=item, is_edit=True)

    item.name = name
    db.session.commit()
    flash("Data master diupdate.", "success")
    return redirect(url_for("masters_list", key=key))


@app.post("/masters/<string:key>/<int:item_id>/delete")
@tenant_required
def masters_delete(key: str, item_id: int):
    cid = current_company_id()
    model, _ = get_master_model(key)
    item = model.query.filter_by(company_id=cid, id=item_id).first_or_404()
    db.session.delete(item)
    db.session.commit()
    flash("Data master dihapus.", "success")
    return redirect(url_for("masters_list", key=key))


# -----------------------
# Export
# -----------------------
@app.get("/export/customers.xlsx")
@tenant_required
def export_customers_xlsx():
    cid = current_company_id()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Customers"

    headers = [
        "ID",
        "Nama",
        "Salesman",
        "Alamat",
        "Telp/WA",
        "Email",
        "PIC",
        "Sumber Prospek",
        "Produk/Jasa Dibutuhkan",
        "Progress",
        "Tgl Prospek",
        "Estimasi Nilai",
        "Rencana Follow Up",
        "Catatan Follow Up Awal",
        "Catatan Follow Up Lanjutan",
        "Komen Manajemen",
    ]
    ws1.append(headers)

    customers = Customer.query.filter_by(company_id=cid).order_by(Customer.id.asc()).all()
    for c in customers:
        ws1.append([
            c.id,
            c.name,
            c.salesman_name or "",
            c.address or "",
            c.phone_wa or "",
            c.email or "",
            c.pic or "",
            c.lead_source.name if c.lead_source else "",
            c.need.name if c.need else "",
            c.progress.name if c.progress else "",
            c.prospect_date.strftime("%Y-%m-%d") if c.prospect_date else "",
            float(c.estimated_value) if c.estimated_value is not None else "",
            c.prospect_next_followup_date.strftime("%Y-%m-%d") if c.prospect_next_followup_date else "",
            c.note_followup_awal or "",
            c.note_followup_lanjutan or "",
            c.management_comment or "",
        ])

    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 22

    ws2 = wb.create_sheet("FollowUps")
    headers2 = ["ID", "Customer ID", "Customer", "Tanggal", "Stage", "Catatan"]
    ws2.append(headers2)

    logs = (
        FollowUpLog.query
        .filter_by(company_id=cid)
        .order_by(FollowUpLog.created_at.desc())
        .all()
    )
    for fu in logs:
        ws2.append([
            fu.id,
            fu.customer_id,
            fu.customer.name if fu.customer else "",
            fu.created_at.strftime("%Y-%m-%d %H:%M"),
            fu.stage.name if fu.stage else "",
            fu.note or "",
        ])

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 24

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="customers_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# -----------------------
# User Management (tetap ada)
# -----------------------
@app.get("/users")
@tenant_required
def users_list():
    cid = current_company_id()
    users = User.query.filter_by(company_id=cid).order_by(User.username.asc()).all()
    return render_template("users_list.html", users=users)


@app.get("/users/new")
@tenant_required
def users_new():
    return render_template("user_form.html", user=None)


@app.post("/users/new")
@tenant_required
def users_create():
    cid = current_company_id()

    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")

    if not username or not password:
        flash("Username dan password wajib diisi.", "error")
        return redirect(url_for("users_new"))

    if User.query.filter_by(username=username).first():
        flash("Username sudah ada.", "error")
        return redirect(url_for("users_new"))

    u = User(username=username, company_id=cid)
    u.set_password(password)
    db.session.add(u)
    db.session.commit()

    flash("User berhasil dibuat.", "success")
    return redirect(url_for("users_list"))


# -----------------------
# Init DB (create tables + schema + seed)
# -----------------------
with app.app_context():
    db.create_all()
    ensure_schema()
    ensure_seed_data()


# -----------------------
# Run local
# -----------------------
if __name__ == "__main__":
    app.run(debug=True)
